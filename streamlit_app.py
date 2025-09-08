import io
import math
import re
import sys
import subprocess
import numpy as np
import pandas as pd
import streamlit as st
from html import escape as _html_escape

# --------------------------------------------------
# Auto-install helper (we try, but if env blocks installs we have XML fallback)
# --------------------------------------------------
def _ensure_pkg(pkg_name, spec=None):
    try:
        __import__(pkg_name)
        return True
    except Exception:
        pass
    try:
        with st.spinner(f"Installing dependency: {pkg_name}…"):
            cmd = [sys.executable, "-m", "pip", "install", spec or pkg_name]
            subprocess.check_call(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        __import__(pkg_name)
        return True
    except Exception:
        return False

@st.cache_resource(show_spinner=False)
def ensure_excel_engine_for_write() -> str:
    if _ensure_pkg("xlsxwriter", "xlsxwriter>=3.2.0"):
        return "xlsxwriter"
    if _ensure_pkg("openpyxl", "openpyxl>=3.1.5"):
        return "openpyxl"
    return ""  # neither available

@st.cache_resource(show_spinner=False)
def ensure_excel_reader():
    _ensure_pkg("openpyxl", "openpyxl>=3.1.5")


# --------------------------------------------------
# Utilities
# --------------------------------------------------
def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    keep_cols = []
    for c in df.columns:
        s = str(c).strip()
        if s and not s.lower().startswith("unnamed"):
            keep_cols.append(c)
    df = df.loc[:, keep_cols]
    df.columns = [re.sub(r"\s+", " ", str(c).strip()) for c in df.columns]
    return df

def is_missing_like(x):
    if pd.isna(x):
        return True
    if isinstance(x, str) and x.strip().lower() in {"", "na", "n/a", "null", "none"}:
        return True
    return False

def parse_timestamp(s):
    if pd.isna(s):
        return pd.NaT
    return pd.to_datetime(s, utc=True, errors="coerce")

def as_int_or_nan(x):
    if is_missing_like(x):
        return np.nan
    try:
        return int(float(str(x).strip()))
    except Exception:
        return np.nan

def is_false_like(value):
    if isinstance(value, bool):
        return value is False
    if is_missing_like(value):
        return False
    if isinstance(value, (int, float)):
        return int(value) == 0
    if isinstance(value, str):
        return value.strip().lower() in {"false", "no", "0"}
    return False

def round_half_up_days(days_float):
    if pd.isna(days_float):
        return np.nan
    return math.floor(days_float + 0.5)

def split_city_state(value):
    if is_missing_like(value):
        return "", ""
    txt = str(value)
    parts = re.split(r"\s*-\s*", txt, maxsplit=1)
    if len(parts) == 1:
        parts = txt.split("-", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return txt.strip(), ""


# --------------------------------------------------
# Robust loader (CSV / Excel)
# --------------------------------------------------
def load_table(uploaded_file) -> pd.DataFrame:
    raw = uploaded_file.read()
    name = (uploaded_file.name or "").lower()

    # Excel by extension or ZIP signature
    if name.endswith((".xlsx", ".xls")) or raw[:2] == b"PK":
        try:
            ensure_excel_reader()
            return pd.read_excel(io.BytesIO(raw))
        except Exception:
            pass

    # Try CSV with multiple encodings + sniffer
    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1", "utf-16", "utf-16-le", "utf-16-be"]
    for enc in encodings:
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc, sep=None, engine="python", on_bad_lines="skip")
        except Exception:
            continue

    # Last resort: attempt Excel again
    ensure_excel_reader()
    return pd.read_excel(io.BytesIO(raw))


# --------------------------------------------------
# RAW → Data mapping (A..U)
# --------------------------------------------------
DATA_COLUMNS_ORDER = [
    "Carrier Name",
    "Bill of Lading",
    "Tracked",
    "Pickup Name",
    "Pickup City State",
    "Pickup Country",
    "Dropoff Name",
    "Dropoff City State",
    "Dropoff Country",
    "Final Status Reason",
    "Pickup Arrival Utc Timestamp Raw",
    "Pickup Departure Utc Timestamp Raw",
    "Dropoff Arrival Utc Timestamp Raw",
    "Dropoff Departure Utc Timestamp Raw",
    "Nb Milestones Expected",
    "Nb Milestones Received",
    "Milestones Achieved Percentage",
    "Latency Updates Received",
    "Latency Updates Passed",
    "Shipment Latency Percentage",
    "Average Latency (min)",
]

def parse_received_expected(series_ratio: pd.Series):
    s = series_ratio.astype(str)
    m = s.str.extract(r"(-?\d+)\s*/\s*(-?\d+)")
    received = pd.to_numeric(m[0], errors="coerce")
    expected = pd.to_numeric(m[1], errors="coerce")
    return received, expected

def build_data_from_raw(df_raw: pd.DataFrame) -> pd.DataFrame:
    df = normalize_headers(df_raw)
    n = len(df)

    def col(name):
        return df.get(name, pd.Series([np.nan] * n))

    rec, exp = parse_received_expected(col("# Of Milestones received / # Of Milestones expected"))
    updates_total = pd.to_numeric(col("# Updates Received"), errors="coerce")
    updates_passed = pd.to_numeric(col("# Updates Received < 10 mins"), errors="coerce")

    with np.errstate(divide="ignore", invalid="ignore"):
        milestones_pct = (rec / exp) * 100.0
        ship_latency_pct = (updates_passed / updates_total) * 100.0

    data = pd.DataFrame({
        "Carrier Name": col("Carrier Name"),
        "Bill of Lading": col("Bill of Lading"),
        "Tracked": col("Tracked"),
        "Pickup Name": col("Pickup Name"),
        "Pickup City State": col("Pickup City State"),
        "Pickup Country": col("Pickup Country"),
        "Dropoff Name": col("Final Destination Name"),
        "Dropoff City State": col("Final Destination City State"),
        "Dropoff Country": col("Final Destination Country"),
        "Final Status Reason": col("Final Status Reason"),
        "Pickup Arrival Utc Timestamp Raw": col("Pickup Arrival Milestone (UTC)"),
        "Pickup Departure Utc Timestamp Raw": col("Pickup Departure Milestone (UTC)"),
        "Dropoff Arrival Utc Timestamp Raw": col("Final Destination Arrival Milestone (UTC)"),
        "Dropoff Departure Utc Timestamp Raw": col("Final Destination Departure Milestone (UTC)"),
        "Nb Milestones Expected": exp,
        "Nb Milestones Received": rec,
        "Milestones Achieved Percentage": milestones_pct,
        "Latency Updates Received": updates_total,
        "Latency Updates Passed": updates_passed,
        "Shipment Latency Percentage": ship_latency_pct,
        "Average Latency (min)": pd.to_numeric(col("Average Latency (min)"), errors="coerce"),
    })

    data = data[DATA_COLUMNS_ORDER]
    return data


# --------------------------------------------------
# V column logic + Summary dataframes
# --------------------------------------------------
def compute_in_transit_time_row(row):
    tracked_val = row.get("Tracked", np.nan)
    nb_recv = row.get("Nb Milestones Received", np.nan)
    milestones_i = as_int_or_nan(nb_recv)
    if is_false_like(tracked_val) or (pd.isna(milestones_i) or milestones_i == 0):
        return "Untracked"

    pick_dep = parse_timestamp(row.get("Pickup Departure Utc Timestamp Raw"))
    drop_arr = parse_timestamp(row.get("Dropoff Arrival Utc Timestamp Raw"))
    if pd.isna(pick_dep) or pd.isna(drop_arr):
        return "Missing Milestone"

    delta_days = (drop_arr - pick_dep).total_seconds() / (24 * 3600)
    if delta_days <= 0:
        return "Missing Milestone"

    return int(round_half_up_days(delta_days))

def build_summary_sheet(df_data):
    v = df_data["In-Transit Time"]
    is_numeric_v = pd.to_numeric(v, errors="coerce").notna()

    count_tracked = int(is_numeric_v.sum())
    count_missing = int((v == "Missing Milestone").sum())
    count_untracked = int((v == "Untracked").sum())
    grand_total = count_tracked + count_missing + count_untracked

    numeric_vals = pd.to_numeric(v, errors="coerce")
    avg_days_all = float(numeric_vals.dropna().mean()) if numeric_vals.notna().any() else np.nan

    df_numeric = df_data[is_numeric_v].copy()

    pick_city, pick_state = zip(*df_numeric["Pickup City State"].map(split_city_state)) if len(df_numeric) else ([], [])
    drop_city, drop_state = zip(*df_numeric["Dropoff City State"].map(split_city_state)) if len(df_numeric) else ([], [])

    summary_main = pd.DataFrame({
        "Bill of Lading": df_numeric["Bill of Lading"].astype(str),
        "Pickup Name": df_numeric["Pickup Name"].astype(str),
        "Pickup City": list(pick_city),
        "Pickup State": list(pick_state),
        "Pickup Country": df_numeric["Pickup Country"].astype(str),
        "Dropoff Name": df_numeric["Dropoff Name"].astype(str),
        "Dropoff City": list(drop_city),
        "Dropoff State": list(drop_state),
        "Dropoff Country": df_numeric["Dropoff Country"].astype(str),
        "Average of In-Transit Time": pd.to_numeric(df_numeric["In-Transit Time"], errors="coerce").astype("Int64"),
    })

    small_counts = {
        "tracked": count_tracked,
        "missed": count_missing,
        "untracked": count_untracked,
        "grand": grand_total,
        "avg": avg_days_all,
    }
    return summary_main, small_counts


# --------------------------------------------------
# Excel writer (tries .xlsx; if engines missing, writes Excel 2003 XML that Excel opens)
# --------------------------------------------------
def write_excel_report(df_data, summary_main, small_counts):
    engine = ensure_excel_engine_for_write()
    if engine == "xlsxwriter":
        return _xlsxwriter_bytes(df_data, summary_main, small_counts), "xlsx"
    if engine == "openpyxl":
        return _openpyxl_bytes(df_data, summary_main, small_counts), "xlsx"
    # Fallback: SpreadsheetML XML (no external deps)
    return _excel_xml_bytes(df_data, summary_main, small_counts), "xml"

def _xlsxwriter_bytes(df_data, summary_main, sc):
    import xlsxwriter  # noqa: F401
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        # Data sheet
        df_data.to_excel(writer, sheet_name="Data", index=False)
        ws_data = writer.sheets["Data"]
        for idx, col in enumerate(df_data.columns):
            width = min(50, max(12, int(df_data[col].astype(str).str.len().quantile(0.9)) + 2))
            ws_data.set_column(idx, idx, width)

        # Summary sheet
        wb = writer.book
        ws = wb.add_worksheet("Summary")

        fmt_blue_bold_border = wb.add_format({"bold": True, "bg_color": "#D9EDF7", "border": 1})
        fmt_bold = wb.add_format({"bold": True})
        fmt_border = wb.add_format({"border": 1})

        # Column widths for A–J
        ws.set_column("A:A", 22)  # Label / Bill of Lading
        ws.set_column("B:B", 18)  # Shipment Count / Pickup Name
        ws.set_column("C:C", 10)  # blank top / Pickup City
        ws.set_column("D:D", 28)  # Average (top) / Pickup State
        ws.set_column("E:E", 34)  # Time taken... / Pickup Country
        ws.set_column("F:J", 22)  # rest

        # ----- Small table A1:E5 -----
        ws.write(0, 0, "Label", fmt_blue_bold_border)
        ws.write(0, 1, "Shipment Count", fmt_blue_bold_border)
        ws.write(0, 2, "", fmt_border)  # blank column
        ws.write(0, 3, "Average of In-Transit Time", fmt_blue_bold_border)
        ws.write(0, 4, "Time taken from Departure to Arrival", fmt_border)

        ws.write(1, 0, "Tracked")
        ws.write(2, 0, "Missed Milestone")
        ws.write(3, 0, "Untracked")
        ws.write(4, 0, "Grand Total", fmt_blue_bold_border)

        ws.write_number(1, 1, int(sc["tracked"]))
        ws.write_number(2, 1, int(sc["missed"]))
        ws.write_number(3, 1, int(sc["untracked"]))
        ws.write_formula(4, 1, "=SUM(B2:B4)", fmt_blue_bold_border)

        if pd.notna(sc["avg"]):
            ws.write_number(4, 3, float(sc["avg"]))
        else:
            ws.write(4, 3, "")

        for r in range(1, 5):
            for c in range(0, 5):
                ws.write_blank(r, c, None, fmt_border)

        # Row 6 blank

        # ----- Main table at row 7 -----
        start_row = 6
        headers = [
            "Bill of Lading", "Pickup Name", "Pickup City", "Pickup State", "Pickup Country",
            "Dropoff Name", "Dropoff City", "Dropoff State", "Dropoff Country",
            "Average of In-Transit Time"
        ]
        for c, h in enumerate(headers):
            ws.write(start_row, c, h, fmt_blue_bold_border)

        for i, (_, row) in enumerate(summary_main.iterrows(), start=1):
            r = start_row + i
            for c_idx, col_name in enumerate(headers):
                val = row[col_name]
                if col_name == "Bill of Lading":
                    ws.write(r, c_idx, "" if pd.isna(val) else str(val), fmt_bold)
                elif col_name == "Average of In-Transit Time":
                    if pd.isna(val):
                        ws.write(r, c_idx, "")
                    else:
                        try:
                            ws.write_number(r, c_idx, float(val))
                        except Exception:
                            ws.write(r, c_idx, str(val))
                else:
                    ws.write(r, c_idx, "" if pd.isna(val) else str(val))

        last_data_row = start_row + len(summary_main) + 1
        ws.write(last_data_row, 0, "Grand Total", fmt_blue_bold_border)
        if len(summary_main) > 0:
            first_j = start_row + 1
            last_j = start_row + len(summary_main)
            ws.write_formula(last_data_row, 9, f"=AVERAGE(J{first_j+1}:J{last_j+1})", fmt_blue_bold_border)
        else:
            ws.write(last_data_row, 9, "", fmt_blue_bold_border)

    output.seek(0)
    return output

def _openpyxl_bytes(df_data, summary_main, sc):
    _ = _ensure_pkg("openpyxl", "openpyxl>=3.1.5")
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Data sheet
        df_data.to_excel(writer, sheet_name="Data", index=False)
        ws_data = writer.sheets["Data"]
        for idx, col in enumerate(df_data.columns, start=1):
            lens = df_data[col].astype(str).str.len()
            q = int(lens.quantile(0.9)) if len(lens) else 12
            ws_data.column_dimensions[get_column_letter(idx)].width = min(50, max(12, q + 2))

        # Summary sheet
        wb = writer.book
        ws = wb.create_sheet("Summary")

        blue_fill = PatternFill(fill_type="solid", start_color="FFD9EDF7", end_color="FFD9EDF7")
        bold_font = Font(bold=True)
        thin = Side(style="thin", color="FF000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def set_cell(r, c, value, blue=False, bold=False, bordered=False):
            cell = ws.cell(row=r, column=c, value=value)
            if blue:
                cell.fill = blue_fill
            if bold:
                cell.font = Font(bold=True)
            if bordered:
                cell.border = border
            return cell

        widths = {1: 22, 2: 18, 3: 10, 4: 28, 5: 34, 6: 22, 7: 22, 8: 18, 9: 18, 10: 22}
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Small table A1:E5
        set_cell(1, 1, "Label", blue=True, bold=True, bordered=True)
        set_cell(1, 2, "Shipment Count", blue=True, bold=True, bordered=True)
        set_cell(1, 3, "", bordered=True)  # blank
        set_cell(1, 4, "Average of In-Transit Time", blue=True, bold=True, bordered=True)
        set_cell(1, 5, "Time taken from Departure to Arrival", bordered=True)

        set_cell(2, 1, "Tracked")
        set_cell(3, 1, "Missed Milestone")
        set_cell(4, 1, "Untracked")
        set_cell(5, 1, "Grand Total", blue=True, bold=True, bordered=True)

        set_cell(2, 2, int(sc["tracked"]))
        set_cell(3, 2, int(sc["missed"]))
        set_cell(4, 2, int(sc["untracked"]))
        b5 = ws.cell(row=5, column=2); b5.value = "=SUM(B2:B4)"; b5.fill = blue_fill; b5.font = bold_font; b5.border = border

        set_cell(5, 4, float(sc["avg"]) if pd.notna(sc["avg"]) else "")

        for r in range(2, 6):
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = border

        # Row 6 blank

        # Main table at row 7
        header_row = 7
        headers = [
            "Bill of Lading", "Pickup Name", "Pickup City", "Pickup State", "Pickup Country",
            "Dropoff Name", "Dropoff City", "Dropoff State", "Dropoff Country",
            "Average of In-Transit Time"
        ]
        for c, h in enumerate(headers, start=1):
            set_cell(header_row, c, h, blue=True, bold=True, bordered=True)

        for i, (_, row) in enumerate(summary_main.iterrows(), start=1):
            r = header_row + i
            for c_idx, col_name in enumerate(headers, start=1):
                val = row[col_name]
                cell = ws.cell(row=r, column=c_idx, value=("" if pd.isna(val) else str(val)))
                if col_name == "Bill of Lading":
                    cell.font = bold_font
                if col_name == "Average of In-Transit Time":
                    if pd.isna(val): cell.value = ""
                    else:
                        try: cell.value = float(val)
                        except Exception: pass

        last_data_row = header_row + len(summary_main) + 1
        set_cell(last_data_row, 1, "Grand Total", blue=True, bold=True, bordered=True)
        if len(summary_main) > 0:
            first_j = header_row + 1; last_j = header_row + len(summary_main)
            jcell = ws.cell(row=last_data_row, column=10)
            jcell.value = f"=AVERAGE(J{first_j}:J{last_j})"
            jcell.fill = blue_fill; jcell.font = bold_font; jcell.border = border
        else:
            set_cell(last_data_row, 10, "", blue=True, bold=True, bordered=True)

    output.seek(0)
    return output

# ---------- SpreadsheetML (Excel 2003 XML) fallback ----------
def _excel_xml_bytes(df_data, summary_main, sc):
    # Simple XML helpers
    def x(s):  # escape text
        return _html_escape("" if s is None else str(s))

    # Styles: Default, BlueBold (with borders), Bold, Border
    styles = """
    <Styles>
      <Style ss:ID="Default" ss:Name="Normal">
        <Alignment ss:Vertical="Bottom"/>
        <Borders/>
        <Font/>
        <Interior/>
        <NumberFormat/>
        <Protection/>
      </Style>
      <Style ss:ID="BlueBold">
        <Font ss:Bold="1"/>
        <Interior ss:Color="#D9EDF7" ss:Pattern="Solid"/>
        <Borders>
          <Border ss:Position="Left" ss:LineStyle="Continuous" ss:Weight="1"/>
          <Border ss:Position="Right" ss:LineStyle="Continuous" ss:Weight="1"/>
          <Border ss:Position="Top" ss:LineStyle="Continuous" ss:Weight="1"/>
          <Border ss:Position="Bottom" ss:LineStyle="Continuous" ss:Weight="1"/>
        </Borders>
      </Style>
      <Style ss:ID="Bold">
        <Font ss:Bold="1"/>
      </Style>
      <Style ss:ID="Border">
        <Borders>
          <Border ss:Position="Left" ss:LineStyle="Continuous" ss:Weight="1"/>
          <Border ss:Position="Right" ss:LineStyle="Continuous" ss:Weight="1"/>
          <Border ss:Position="Top" ss:LineStyle="Continuous" ss:Weight="1"/>
          <Border ss:Position="Bottom" ss:LineStyle="Continuous" ss:Weight="1"/>
        </Borders>
      </Style>
    </Styles>
    """

    def xml_cell(value, style=None, num=False):
        style_attr = f' ss:StyleID="{style}"' if style else ""
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return f'<Cell{style_attr}/>'
        if num:
            return f'<Cell{style_attr}><Data ss:Type="Number">{value}</Data></Cell>'
        else:
            return f'<Cell{style_attr}><Data ss:Type="String">{x(value)}</Data></Cell>'

    # Worksheet: Data
    def sheet_data():
        rows = []
        # header
        hdr = "".join(xml_cell(col) for col in df_data.columns)
        rows.append(f"<Row>{hdr}</Row>")
        # data
        for _, r in df_data.iterrows():
            cells = []
            for col in df_data.columns:
                v = r[col]
                # try numeric where appropriate
                if pd.api.types.is_numeric_dtype(df_data[col]):
                    cells.append(xml_cell("" if pd.isna(v) else float(v), num=True))
                else:
                    cells.append(xml_cell("" if pd.isna(v) else v))
            rows.append(f"<Row>{''.join(cells)}</Row>")
        return f"<Table>{''.join(rows)}</Table>"

    # Worksheet: Summary (top + main)
    def sheet_summary():
        rows = []

        # A1:E1 headers
        row1 = "".join([
            xml_cell("Label", style="BlueBold"),
            xml_cell("Shipment Count", style="BlueBold"),
            xml_cell("", style="Border"),
            xml_cell("Average of In-Transit Time", style="BlueBold"),
            xml_cell("Time taken from Departure to Arrival", style="Border"),
        ])
        rows.append(f"<Row>{row1}</Row>")

        # A2..A4 + counts in B
        rows.append(f"<Row>{xml_cell('Tracked')}{xml_cell(sc['tracked'], num=True)}{xml_cell('', 'Border')}{xml_cell('')}{xml_cell('', 'Border')}</Row>")
        rows.append(f"<Row>{xml_cell('Missed Milestone')}{xml_cell(sc['missed'], num=True)}{xml_cell('', 'Border')}{xml_cell('')}{xml_cell('', 'Border')}</Row>")
        rows.append(f"<Row>{xml_cell('Untracked')}{xml_cell(sc['untracked'], num=True)}{xml_cell('', 'Border')}{xml_cell('')}{xml_cell('', 'Border')}</Row>")

        # A5/B5 blue+bold; B5 = sum (we'll compute since formulas are verbose)
        b5 = (sc['tracked'] or 0) + (sc['missed'] or 0) + (sc['untracked'] or 0)
        d5 = float(sc["avg"]) if pd.notna(sc["avg"]) else ""

        rows.append(f"<Row>"
                    f"{xml_cell('Grand Total', style='BlueBold')}"
                    f"{xml_cell(b5, style='BlueBold', num=True)}"
                    f"{xml_cell('', 'Border')}"
                    f"{xml_cell(d5, num=isinstance(d5, (int,float)))}"
                    f"{xml_cell('', 'Border')}"
                    f"</Row>")

        # Row 6 blank
        rows.append("<Row></Row>")

        # Row 7 header A..J (BlueBold)
        headers = [
            "Bill of Lading", "Pickup Name", "Pickup City", "Pickup State", "Pickup Country",
            "Dropoff Name", "Dropoff City", "Dropoff State", "Dropoff Country",
            "Average of In-Transit Time"
        ]
        rows.append("<Row>" + "".join(xml_cell(h, style="BlueBold") for h in headers) + "</Row>")

        # Data rows
        for _, r in summary_main.iterrows():
            cells = []
            # Bold for Bill of Lading
            cells.append(xml_cell("" if pd.isna(r["Bill of Lading"]) else r["Bill of Lading"], style="Bold"))
            cells.append(xml_cell("" if pd.isna(r["Pickup Name"]) else r["Pickup Name"]))
            cells.append(xml_cell("" if pd.isna(r["Pickup City"]) else r["Pickup City"]))
            cells.append(xml_cell("" if pd.isna(r["Pickup State"]) else r["Pickup State"]))
            cells.append(xml_cell("" if pd.isna(r["Pickup Country"]) else r["Pickup Country"]))
            cells.append(xml_cell("" if pd.isna(r["Dropoff Name"]) else r["Dropoff Name"]))
            cells.append(xml_cell("" if pd.isna(r["Dropoff City"]) else r["Dropoff City"]))
            cells.append(xml_cell("" if pd.isna(r["Dropoff State"]) else r["Dropoff State"]))
            cells.append(xml_cell("" if pd.isna(r["Dropoff Country"]) else r["Dropoff Country"]))
            av = r["Average of In-Transit Time"]
            if pd.isna(av):
                cells.append(xml_cell(""))
            else:
                try:
                    cells.append(xml_cell(float(av), num=True))
                except Exception:
                    cells.append(xml_cell(str(av)))
            rows.append(f"<Row>{''.join(cells)}</Row>")

        # Grand total average row
        if len(summary_main) > 0:
            avg_val = pd.to_numeric(summary_main["Average of In-Transit Time"], errors="coerce").dropna().astype(float)
            javg = float(avg_val.mean()) if len(avg_val) else ""
        else:
            javg = ""
        rows.append("<Row>" +
                    xml_cell("Grand Total", style="BlueBold") +
                    "".join(xml_cell("") for _ in range(8)) +
                    (xml_cell(javg, style="BlueBold", num=isinstance(javg, (int,float)))) +
                    "</Row>")

        return f"<Table>{''.join(rows)}</Table>"

    xml = f"""<?xml version="1.0"?>
<?mso-application progid="Excel.Sheet"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:html="http://www.w3.org/TR/REC-html40">
{styles}
<Worksheet ss:Name="Data">
{sheet_data()}
</Worksheet>
<Worksheet ss:Name="Summary">
{sheet_summary()}
</Worksheet>
</Workbook>
"""
    return io.BytesIO(xml.encode("utf-8"))

# --------------------------------------------------
# Streamlit UI & flow
# --------------------------------------------------
st.set_page_config(page_title="FTL In-Transit Builder", page_icon="🚚", layout="wide")
st.title("FTL In-Transit Time Processor (RAW → Data → Summary)")

mode_col, info_col = st.columns([1, 3])
with mode_col:
    mode = st.selectbox("Mode", options=["FTL"], index=0)
with info_col:
    eng = ensure_excel_engine_for_write()
    engine_msg = eng if eng else "none (will export Excel XML fallback)"
    st.caption("Upload RAW CSV/XLSX. We build the Data sheet (A..U), add V=In-Transit Time "
               "(round: <.5 down, ≥.5 up), then generate the Summary sheet. "
               f"Excel writer engine: **{engine_msg}**")

uploaded = st.file_uploader("Upload RAW CSV or Excel", type=["csv", "xlsx", "xls"])

if uploaded is not None:
    try:
        df_raw = load_table(uploaded)
        df_data = build_data_from_raw(df_raw)

        # Compute V on the Data sheet
        df_data["In-Transit Time"] = df_data.apply(compute_in_transit_time_row, axis=1)

        # Build Summary parts
        summary_main, small_counts = build_summary_sheet(df_data)

        # Build single Excel report
        blob, ext = write_excel_report(df_data, summary_main, small_counts)
        fname = "FTL_Data_and_Summary.xlsx" if ext == "xlsx" else "FTL_Data_and_Summary.xml"

        st.success("Processed! Click below to download your Excel report.")
        st.download_button(
            "⬇️ Download Report (Excel)",
            data=blob,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext == "xlsx" else "application/vnd.ms-excel",
            use_container_width=True
        )

        # Optional previews
        with st.expander("Preview: Data (A..U + V)"):
            st.dataframe(df_data.head(50))
        with st.expander("Preview: Summary main table"):
            st.dataframe(summary_main.head(50))

    except Exception as e:
        st.error(f"Could not process this file. Details: {e}")
else:
    st.info("Upload your raw file to begin.")
